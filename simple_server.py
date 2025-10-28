"""
Complete Server - Issues + 3D Viewer
Final version with all endpoints
Run this file only: python simple_server.py
"""

from flask import Flask, jsonify, send_file, request
from flask_cors import CORS
import os
import base64
import time
from dotenv import load_dotenv
import json

load_dotenv()

app = Flask(__name__)
CORS(app)

# Configuration
CLIENT_ID = os.getenv("APS_CLIENT_ID", "").strip()
CLIENT_SECRET = os.getenv("APS_CLIENT_SECRET", "").strip()
VERSION_URN = os.getenv("VERSION_URN", "").strip()
BASE_URL = "https://developer.api.autodesk.com"

# ============= MISSING VARIABLES - NOW ADDED =============
# Token cache for 3D viewer (2-legged OAuth)
token_cache = {
    'token': None,
    'expires_at': 0
}

# Issues cache (to avoid fetching too often)
issues_cache = {
    'data': None,
    'timestamp': 0
}

# Cache duration: 5 minutes
CACHE_DURATION = 300
# =========================================================

# Model URN mapping cache
MODEL_URN_CACHE = {}

# Import issues fetcher
FETCHER_AVAILABLE = False
fetch_all_issues = None

try:
    from acc_issues_fetcher_simple import fetch_all_issues
    FETCHER_AVAILABLE = True
    print("✓ Issues fetcher imported successfully")
except ImportError as e:
    print(f"⚠ Issues fetcher not available: {e}")
except Exception as e:
    print(f"⚠ Error importing fetcher: {e}")


def get_access_token():
    """Get access token for 3D viewer (2-legged OAuth)"""
    if token_cache['token'] and time.time() < token_cache['expires_at']:
        return token_cache['token']
    
    import requests
    url = f"{BASE_URL}/authentication/v2/token"
    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
        "scope": "data:read viewables:read"
    }
    
    response = requests.post(
        url,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data=data,
        timeout=30
    )
    
    if response.status_code == 200:
        result = response.json()
        token_cache['token'] = result['access_token']
        token_cache['expires_at'] = time.time() + result.get('expires_in', 3600) - 60
        return result['access_token']
    
    return None

def build_model_urn_mapping():
    """Build mapping from viewable_guid to model URN"""
    global MODEL_URN_CACHE
    
    if not issues_cache['data']:
        return {}
    
    # Get unique viewables from issues
    unique_viewables = {}
    for issue in issues_cache['data']:
        vg = issue.get('viewable_guid')
        vn = issue.get('viewable_name', 'Model')
        if vg and vg not in unique_viewables:
            unique_viewables[vg] = vn
    
    print(f"\n📊 Found {len(unique_viewables)} unique models:")
    for vg, vn in unique_viewables.items():
        print(f"   - {vn}")
    
    # Get URNs from environment variables
    hofuf_urn = os.getenv("HOFUF_URN", "").strip()
    snowdon_urn = os.getenv("SNOWDON_STR_URN", "").strip()
    
    if hofuf_urn and snowdon_urn:
        print("✅ Using URNs from .env file")
        MODEL_URN_CACHE = {
            'd039209d-a250-1473-1dd9-a3953b7c2e9b': hofuf_urn,
            '5fbd4c90-ff9e-87ff-78ff-b3654b67f6e4': snowdon_urn
        }
    else:
        print("⚠️ Add HOFUF_URN and SNOWDON_STR_URN to .env file")
    
    return MODEL_URN_CACHE

@app.route('/')
def index():
    """Serve 3D viewer HTML"""
    try:
        html_file = os.path.join(os.path.dirname(__file__), 'forge_viewer_powerbi.html')
        if os.path.exists(html_file):
            return send_file(html_file)
        return jsonify({'error': 'Viewer HTML not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/token', methods=['GET', 'POST'])
def get_token_endpoint():
    """Token for 3D viewer (2-legged OAuth for viewing)"""
    try:
        if token_cache['token'] and time.time() < token_cache['expires_at']:
            remaining = int(token_cache['expires_at'] - time.time())
            return jsonify({
                'access_token': token_cache['token'],
                'expires_in': remaining
            })
        
        token = get_access_token()
        if token:
            remaining = int(token_cache['expires_at'] - time.time())
            return jsonify({
                'access_token': token,
                'expires_in': remaining
            })
        return jsonify({'error': 'Failed to get token'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/model-urn', methods=['GET'])
def get_model_urn():
    """Return model URN"""
    try:
        if not VERSION_URN:
            return jsonify({'error': 'VERSION_URN not set'}), 500
        
        urn_encoded = base64.urlsafe_b64encode(
            VERSION_URN.encode('utf-8')
        ).decode('utf-8').rstrip('=')
        
        return jsonify({
            'urn': urn_encoded,
            'version_urn': VERSION_URN
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/issues')
def api_issues():
    """MAIN ENDPOINT - Get all issues for Power BI"""
    print("\n📋 /api/issues called")
    
    try:
        # Check cache
        should_refresh = (
            issues_cache['data'] is None or 
            time.time() - issues_cache['timestamp'] > CACHE_DURATION
        )
        
        if should_refresh:
            print("   Fetching fresh issues...")
            
            if not FETCHER_AVAILABLE or not fetch_all_issues:
                print("   ❌ Fetcher not available")
                return jsonify({
                    'error': 'Issues fetcher not available',
                    'message': 'Check acc_issues_fetcher_simple.py'
                }), 500
            
            try:
                issues_data = fetch_all_issues()
                issues_cache['data'] = issues_data
                issues_cache['timestamp'] = time.time()
                print(f"   ✅ Got {len(issues_data)} issues")
            except Exception as e:
                print(f"   ❌ Fetch failed: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({
                    'error': str(e),
                    'message': 'Failed to fetch issues'
                }), 500
        else:
            age = int(time.time() - issues_cache['timestamp'])
            print(f"   Using cache ({len(issues_cache['data'])} issues, {age}s old)")
        
        return jsonify(issues_cache['data'])
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/issues/stats')
def api_stats():
    """Get issue statistics"""
    try:
        if not issues_cache['data']:
            api_issues()
        
        issues = issues_cache['data'] or []
        return jsonify({
            'total': len(issues),
            'open': len([i for i in issues if i.get('status', '').lower() == 'open']),
            'closed': len([i for i in issues if i.get('status', '').lower() == 'closed']),
            'high': len([i for i in issues if i.get('severity', '').lower() == 'high']),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/model-urn-for-viewable')
def get_model_urn_for_viewable():
    """Get the model URN for a specific viewable_guid"""
    try:
        viewable_guid = request.args.get('viewable_guid')
        
        if not viewable_guid:
            return jsonify({'error': 'viewable_guid required'}), 400
        
        # Build mapping if not done
        if not MODEL_URN_CACHE:
            build_model_urn_mapping()
        
        # Get the URN
        model_urn = MODEL_URN_CACHE.get(viewable_guid)
        
        if not model_urn:
            print(f"   ❌ No URN found for viewable: {viewable_guid}")
            return jsonify({'error': 'URN not found for this viewable'}), 404
        
        # Get viewable name
        viewable_name = 'Model'
        if issues_cache['data']:
            for issue in issues_cache['data']:
                if issue.get('viewable_guid') == viewable_guid:
                    viewable_name = issue.get('viewable_name', 'Model')
                    if viewable_name == '{3D}':
                        viewable_name = 'Snowdon Structure'
                    if '.' in viewable_name:
                        viewable_name = viewable_name.rsplit('.', 1)[0]
                    break
        
        # Encode URN
        urn_encoded = base64.urlsafe_b64encode(
            model_urn.encode('utf-8')
        ).decode('utf-8').rstrip('=')
        
        print(f"   ✅ Returning URN for {viewable_name}")
        
        return jsonify({
            'urn': urn_encoded,
            'viewable_guid': viewable_guid,
            'viewable_name': viewable_name,
            'model_name': viewable_name
        })
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/thumbnail-table.html')
def thumbnail_table():
    """HTML table with clickable thumbnails - FIXED VERSION"""
    print("\n🖼️ /thumbnail-table.html called")
    
    try:
        if not issues_cache['data']:
            if FETCHER_AVAILABLE and fetch_all_issues:
                issues_cache['data'] = fetch_all_issues()
        
        issues = issues_cache['data'] or []
        issues_with_coords = [i for i in issues if i.get('pin_x') and i.get('pin_y') and i.get('pin_z')]
        
        print(f"   📍 Found {len(issues_with_coords)} issues with coordinates")
        
        # Start HTML with embedded styles
        html = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', sans-serif;
            background: #f5f5f5;
            padding: 2px;
        }
        .thumbnail-table { 
            width: 100%; 
            border-collapse: collapse;
            background: white;
        }
        .thumbnail-table thead {
            background: #004E43;
            color: white;
            position: sticky;
            top: 0;
            z-index: 100;
        }
        .thumbnail-table th {
            padding: 1px 1px;
            text-align: center;
            font-size: 12px;
            font-weight: 600;
            top: 0;
            background: #004E43;;
            z-index: 100;
        }
        .thumbnail-table td {
            padding: 10px;
            border-bottom: 1px solid #e0e0e0;
            font-size: 11px;
        }
        .thumbnail-img {
            width: 70px;
            height: 52px;
            object-fit: cover;
            border-radius: 4px;
            cursor: pointer;
            transition: all 0.2s;
            border: 2px solid #ddd;
        }
        .thumbnail-img:hover {
            transform: scale(1.1);
            box-shadow: 0 4px 12px rgba(0,0,0,0.3);
            border-color: #667eea;
        }
        .clickable-row {
            cursor: pointer;
            transition: background 0.15s;
        }
        .clickable-row:hover {
            background: #f8f9fa;
        }
        .status-badge {
            display: inline-block;
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 10px;
            font-weight: 600;
        }
        .status-open { background: #fff3cd; color: #856404; }
        .status-closed { background: #d4edda; color: #155724; }
        
        #debug-log {
            position: fixed;
            bottom: 10px;
            right: 10px;
            background: rgba(0,0,0,0.8);
            color: #0f0;
            padding: 8px;
            border-radius: 4px;
            font-family: monospace;
            font-size: 9px;
            max-width: 250px;
            max-height: 150px;
            overflow-y: auto;
            z-index: 10000;
        }
         /* ========== ADD THESE FILTER STYLES ========== */
        .filter-container {
            background: white;
            padding: 15px;
            margin-bottom: 10px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            display: flex;
            gap: 15px;
            align-items: center;
            flex-wrap: wrap;
        }
        
        .filter-group {
            display: flex;
            flex-direction: column;
            gap: 5px;
        }
        
        .filter-group label {
            font-size: 11px;
            font-weight: 600;
            color: #666;
            text-transform: uppercase;
        }
        
        .filter-group select {
            padding: 8px 12px;
            border: 2px solid #e0e0e0;
            border-radius: 4px;
            font-size: 12px;
            background: white;
            cursor: pointer;
            min-width: 150px;
            transition: border-color 0.2s;
        }
        
        .filter-group select:hover {
            border-color: #667eea;
        }
        
        .filter-group select:focus {
            outline: none;
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }
        
        .clear-filters-btn {
            padding: 8px 16px;
            background: #667eea;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 600;
            transition: background 0.2s;
            margin-top: 18px;
        }
        
        .clear-filters-btn:hover {
            background: #5568d3;
        }
        
        .filter-count {
            margin-top: 18px;
            padding: 8px 12px;
            background: #f5f5f5;
            border-radius: 4px;
            font-size: 12px;
            color: #666;
            font-weight: 600;}
        /* Filter status bar */
        .filter-status-bar {
            background: #f0f0f0;
            padding: 10px;
            margin-bottom: 10px;
            border-radius: 4px;
            font-size: 12px;
            display: none;
        }

        .filter-status-bar.active {
            display: block;
        }

        .filter-tag {
            display: inline-block;
            background: #667eea;
            color: white;
            padding: 4px 8px;
            border-radius: 12px;
            margin-right: 8px;
            font-size: 11px;
        }

        .filter-tag .remove {
            margin-left: 6px;
            cursor: pointer;
            font-weight: bold;
        }
        
        /* Excel-style filter headers */
    .filterable-header {
        position: relative;
        cursor: pointer;
        user-select: none;
    }

    .filterable-header:hover {
        background: linear-gradient(135deg, #5568d3 0%, #6a4a9e 100%);
    }

    .filter-icon {
        font-size: 10px;
        margin-left: 5px;
        opacity: 0.7;
    }

    .filterable-header.filtered .filter-icon {
        color: #ffd700;
        opacity: 1;
        font-weight: bold;
    }

    /* Filter dropdown popup */
    .filter-dropdown {
        position: absolute;
        top: 100%;
        left: 0;
        background: white;
        border: 2px solid #667eea;
        border-radius: 4px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        z-index: 1000;
        min-width: 200px;
        max-height: 300px;
        overflow-y: auto;
        display: none;
        color: #333;  /* ← Dark text */
    }

    .filter-dropdown.active {
        display: block;
    }

    .filter-search {
        width: calc(100% - 20px);
        padding: 8px;
        margin: 10px;
        border: 1px solid #ddd;
        border-radius: 4px;
        font-size: 12px;
        color: #333;  /* ← Dark text */
    }

    .filter-options {
        max-height: 200px;
        overflow-y: auto;
    }

    .filter-option {
        padding: 8px 12px;
        cursor: pointer;
        font-size: 12px;
        display: flex;
        align-items: center;
        gap: 8px;
        color: #333;  /* ← Dark text */
    }

    .filter-option:hover {
        background: #f0f0f0;
    }

    .filter-option label {
        color: #333;  /* ← Dark text */
        cursor: pointer;
        user-select: none;
    }

    .filter-option input[type="checkbox"] {
        cursor: pointer;
    }
            
    </style>
</head>
<body>

<div style="padding: 10px; background: white; margin-bottom: 10px; border-radius: 4px;">
    <!-- Open in Browser link - shown in Power BI -->
    <div id="powerbi-buttons" style="display: none;">
        <a href="http://localhost:5000/thumbnail-table.html" target="_blank" style="
            display: inline-block;
            background: #004E43;
            color: white;
            border: none;
            padding: 5px 5px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 600;
            text-decoration: none;
            box-shadow: 0 2px 4px rgba(0,0,0,0.2);
        ">🌐 Open in Browser</a>
        <span style="font-size: 12px; color: #666; margin-left: 2px;margin-top: 2px; margin-bottom: 2px; display: inline-block;">
            Click to open in external browser for export functionality
        </span>
    </div>
    
    <!-- Export buttons - shown in regular browser -->
    <div id="browser-buttons" style="display: none; gap: 10px; flex-wrap: wrap;">
                
        <button onclick="exportToCSV()" style="
            background: #004E43;
            color: white;
            border: none;
            padding: 5px 5px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 600;
        ">📄 Export CSV (no images)</button>
        
        <button onclick="exportRealExcel()" style="
            background: #27ae60;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 600;
        ">📊 Export Real Excel (with images)</button>
    </div>
</div>

<script>
    // Enhanced detection with debug info
    function detectEnvironment() {
        const debugDiv = document.getElementById('debug-info');
        let debugInfo = '';
        
        try {
            const inIframe = window.self !== window.top;
            const windowLocation = window.location.href;
            const parentAccessible = (function() {
                try {
                    return window.parent.location.href !== window.location.href;
                } catch(e) {
                    return true; // Can't access parent = in iframe
                }
            })();
            
            debugInfo += 'In iframe: ' + inIframe + '<br>';
            debugInfo += 'Parent accessible: ' + parentAccessible + '<br>';
            debugInfo += 'Current URL: ' + windowLocation + '<br>';
            
            if (inIframe || parentAccessible) {
                document.getElementById('powerbi-buttons').style.display = 'block';
                document.getElementById('browser-buttons').style.display = 'none';
                debugInfo += '<strong style="color: red;">MODE: Power BI (iframe detected)</strong>';
            } else {
                document.getElementById('powerbi-buttons').style.display = 'none';
                document.getElementById('browser-buttons').style.display = 'flex';
                debugInfo += '<strong style="color: green;">MODE: Browser (standalone)</strong>';
            }
        } catch (e) {
            debugInfo += '<strong style="color: orange;">ERROR: ' + e.message + '</strong><br>';
            document.getElementById('powerbi-buttons').style.display = 'block';
            document.getElementById('browser-buttons').style.display = 'none';
        }
        
        debugDiv.innerHTML = debugInfo;
    }
    
    detectEnvironment();
</script>





<table class="thumbnail-table">
    <div id="debug-log" style="display:none;">Loading...</div>
     <!-- ========== ADD FILTER CONTAINER HERE ========== -->
    
    <table class="thumbnail-table">
        <thead>
            <tr>
                <th>Thumbnail</th>
                <th class="filterable-header" data-column="display_id">
                    ID <span class="filter-icon">▼</span>
                </th>
                <th class="filterable-header" data-column="title">
                    Title <span class="filter-icon">▼</span>
                </th>
                <th class="filterable-header" data-column="status">
                    Status <span class="filter-icon">▼</span>
                </th>
                <th class="filterable-header" data-column="severity">
                    Severity <span class="filter-icon">▼</span>
                </th>
                <th class="filterable-header" data-column="assigned_to">
                    Assigned To <span class="filter-icon">▼</span>
                </th>
                <th>Comments</th>
                <th>Comments By</th>
            </tr>
        </thead>
        <tbody>
"""
        
        # Add each row
        for idx, issue in enumerate(issues_with_coords):
            issue_id = issue.get('issue_id', '')
            display_id = issue.get('display_id', '')
            title = issue.get('title', 'Untitled')
            status = issue.get('status', 'Unknown')
            severity = issue.get('severity', 'N/A')
            assigned_to = issue.get('assigned_to', 'Unassigned')
            comments = issue.get('comment_1', '')
            comments_by = issue.get('comment_1_by', '')
            
            status_class = 'status-open'
            if 'closed' in status.lower():
                status_class = 'status-closed'
            
            img_id = f"img_{idx}"
            
            html += f"""
        <tr class="clickable-row" onclick="handleRowClick({idx})">
            <td>
                <img id="{img_id}" class="thumbnail-img" alt="{display_id}" />
            </td>
            <td><strong>{display_id}</strong></td>
            <td>{title}</td>
            <td><span class="status-badge {status_class}">{status}</span></td>
            <td>{severity}</td>
            <td>{assigned_to}</td>
            <td>{comments[:50] if comments else 'No comments'}</td>
            <td>{comments_by}</td>
        </tr>
"""
        
        html += """
        </tbody>
    </table>
    
    <script>
        // ========== DEBUG TEST ==========
        alert('TABLE JAVASCRIPT IS RUNNING!');
        console.log('🟢 TABLE SCRIPT STARTED');
        // ================================
        
        const debugLog = document.getElementById('debug-log');
        
        function logDebug(msg) {
            console.log(msg);
            debugLog.innerHTML += msg + '<br>';
            debugLog.scrollTop = debugLog.scrollHeight;
        }
        
        // Store issue data
        const issuesData = [
"""
        
        # Add issue data as JavaScript array


        for issue in issues_with_coords:
            issue_id = issue.get('issue_id', '')
            title = issue.get('title', '')
            thumbnail = issue.get('thumbnail_base64', '')
            
            # Get viewable info
            viewable_name = issue.get('viewable_name', 'Model')
            viewable_guid = issue.get('viewable_guid', '')
            if '.' in viewable_name:
                viewable_name = viewable_name.rsplit('.', 1)[0]
            
            # Properly escape all strings for JavaScript
            issue_id_safe = json.dumps(issue_id)
            display_id_safe = json.dumps(issue.get('display_id', ''))
            title_safe = json.dumps(title)
            status_safe = json.dumps(issue.get('status', ''))
            severity_safe = json.dumps(issue.get('severity', ''))
            assigned_to_safe = json.dumps(issue.get('assigned_to', ''))
            viewable_name_safe = json.dumps(viewable_name)
            viewable_guid_safe = json.dumps(viewable_guid)
            thumbnail_safe = json.dumps(thumbnail)
            comment_1_safe = json.dumps(issue.get('comment_1', ''))
            comment_1_by_safe = json.dumps(issue.get('comment_1_by', ''))
            comment_count = issue.get('comment_count', 0)
            
            html += f"""
                {{
                    issue_id: {issue_id_safe},
                    display_id: {display_id_safe},
                    title: {title_safe},
                    status: {status_safe},
                    severity: {severity_safe},
                    assigned_to: {assigned_to_safe},
                    pin_x: {issue.get('pin_x', 0)},
                    pin_y: {issue.get('pin_y', 0)},
                    pin_z: {issue.get('pin_z', 0)},
                    viewable_name: {viewable_name_safe},
                    viewable_guid: {viewable_guid_safe},
                    thumbnail: {thumbnail_safe},
                    comment_1: {comment_1_safe},
                    comment_1_by: {comment_1_by_safe},
                    comment_count: {comment_count}
                }},
        """
        
        html += """
        ];
        
        // Load images after page loads
        window.onload = function(){
            console.log('🔍 Total issues:', issuesData.length);
            logDebug('Page loaded');            
            issuesData.forEach((issue, idx) => {
                const img = document.getElementById('img_' + idx);
                
                // Debug each image
                console.log(`Issue ${idx} (${issue.display_id}):`, {
                    has_img: !!img,
                    has_thumb: !!issue.thumbnail,
                    thumb_len: issue.thumbnail ? issue.thumbnail.length : 0
                });
                
                if (img) {
                    if (issue.thumbnail && issue.thumbnail.length > 0) {
                        img.src = issue.thumbnail;
                        
                        // Add error handler
                        img.onerror = function() {
                            console.error('❌ Failed:', issue.display_id);
                            this.src = 'data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iNzAiIGhlaWdodD0iNTIiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+PHJlY3Qgd2lkdGg9IjcwIiBoZWlnaHQ9IjUyIiBmaWxsPSIjZTBlMGUwIi8+PHRleHQgeD0iNTAlIiB5PSI1MCUiIGZvbnQtZmFtaWx5PSJBcmlhbCIgZm9udC1zaXplPSIxMCIgZmlsbD0iIzk5OSIgdGV4dC1hbmNob3I9Im1pZGRsZSIgZHk9Ii4zZW0iPk5vIEltYWdlPC90ZXh0Pjwvc3ZnPg==';
                        };
                        
                        img.onload = function() {
                            console.log('✅ Loaded:', issue.display_id);
                        };
                        
                        logDebug('Set img ' + idx);
                    } else {
                        console.warn('⚠️ Empty thumbnail:', issue.display_id);
                        img.src = 'data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iNzAiIGhlaWdodD0iNTIiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+PHJlY3Qgd2lkdGg9IjcwIiBoZWlnaHQ9IjUyIiBmaWxsPSIjZTBlMGUwIi8+PHRleHQgeD0iNTAlIiB5PSI1MCUiIGZvbnQtZmFtaWx5PSJBcmlhbCIgZm9udC1zaXplPSIxMCIgZmlsbD0iIzk5OSIgdGV4dC1hbmNob3I9Im1pZGRsZSIgZHk9Ii4zZW0iPk5vIEltYWdlPC90ZXh0Pjwvc3ZnPg==';
                    }
                }
            });
                    
            logDebug('Done!');
};
                            
        function handleRowClick(idx) {
            const issue = issuesData[idx];
            logDebug('Clicked: ' + issue.display_id);
            sendMessageToViewer(issue);
        }
        
        function sendMessageToViewer(issue) {
            console.log('🔵 CLICK DETECTED:', issue.display_id);
            console.log('   Viewable GUID:', issue.viewable_guid);
            console.log('   Viewable Name:', issue.viewable_name);
            logDebug('Sending message...');
            
            const message = {
                type: 'LOAD_MODEL_AND_NAVIGATE',
                issue_id: issue.issue_id,
                display_id: issue.display_id,
                pin_x: issue.pin_x,
                pin_y: issue.pin_y,
                pin_z: issue.pin_z,
                title: issue.title,
                viewable_name: issue.viewable_name,
                viewable_guid: issue.viewable_guid,
                timestamp: Date.now()
            };
            
            console.log('📤 Sending message:', message);
            
            // Method 1: Post to parent
            try {
                parent.postMessage(message, '*');
                console.log('✅ Sent to parent');
                logDebug('Sent to parent');
            } catch(e) {
                console.error('❌ Parent failed:', e);
                logDebug('Parent failed: ' + e.message);
            }
            
            // Method 2: Post to top window
            try {
                window.top.postMessage(message, '*');
                logDebug('Sent to top');
            } catch(e) {
                logDebug('Top failed: ' + e.message);
            }
            
            // Method 3: Post to opener
            if (window.opener) {
                try {
                    window.opener.postMessage(message, '*');
                    logDebug('Sent to opener');
                } catch(e) {
                    logDebug('Opener failed: ' + e.message);
                }
            }
            
            logDebug('Message broadcast complete');
        }
                              
            function exportRealExcel() {
                console.log('📊 Downloading real Excel...');
                    window.location.href = '/api/export-excel-with-images';
                }
        

        // ========== EXCEL-STYLE FILTER FUNCTIONS ==========
        let activeFilters = {};
        let currentDropdown = null;

        function initColumnFilters() {
            const headers = document.querySelectorAll('.filterable-header');
            
            headers.forEach(header => {
                header.addEventListener('click', function(e) {
                    e.stopPropagation();
                    const column = this.dataset.column;
                    toggleFilterDropdown(this, column);
                });
            });
            
            document.addEventListener('click', function() {
                if (currentDropdown) {
                    currentDropdown.remove();
                    currentDropdown = null;
                }
            });
            
            logDebug('Excel filters initialized');
        }

        function toggleFilterDropdown(headerElement, column) {
            if (currentDropdown) {
                currentDropdown.remove();
                currentDropdown = null;
            }
            
            const values = [...new Set(issuesData.map(item => item[column]))].filter(Boolean).sort();
            
            const dropdown = document.createElement('div');
            dropdown.className = 'filter-dropdown active';
            dropdown.onclick = (e) => e.stopPropagation();
            
            dropdown.innerHTML = `
                <input type="text" class="filter-search" placeholder="Search..." onkeyup="filterDropdownOptions(this)">
                <div class="filter-options">
                    <div class="filter-option">
                        <input type="checkbox" id="select-all-${column}" checked onchange="toggleSelectAll('${column}')">
                        <label for="select-all-${column}"><strong>(Select All)</strong></label>
                    </div>
                    ${values.map(value => `
                        <div class="filter-option" data-value="${value}">
                            <input type="checkbox" id="filter-${column}-${value}" value="${value}" checked>
                            <label for="filter-${column}-${value}">${value}</label>
                        </div>
                    `).join('')}
                </div>
                <div class="filter-actions">
                    <button class="filter-btn filter-btn-apply" onclick="applyColumnFilter('${column}')">OK</button>
                    <button class="filter-btn filter-btn-clear" onclick="clearColumnFilter('${column}')">Clear</button>
                </div>
            `;
            
            headerElement.appendChild(dropdown);
            currentDropdown = dropdown;
            
            if (activeFilters[column]) {
                const checkboxes = dropdown.querySelectorAll('input[type="checkbox"]:not(#select-all-' + column + ')');
                checkboxes.forEach(cb => {
                    cb.checked = activeFilters[column].includes(cb.value);
                });
                updateSelectAll(column);
            }
        }

        function filterDropdownOptions(searchInput) {
            const searchTerm = searchInput.value.toLowerCase();
            const options = searchInput.parentElement.querySelectorAll('.filter-option:not(:first-child)');
            
            options.forEach(option => {
                const text = option.textContent.toLowerCase();
                option.style.display = text.includes(searchTerm) ? 'flex' : 'none';
            });
        }

        function toggleSelectAll(column) {
            const selectAll = document.getElementById('select-all-' + column);
            const checkboxes = document.querySelectorAll(`input[id^="filter-${column}-"]`);
            
            checkboxes.forEach(cb => {
                cb.checked = selectAll.checked;
            });
        }

        function updateSelectAll(column) {
            const selectAll = document.getElementById('select-all-' + column);
            const checkboxes = document.querySelectorAll(`input[id^="filter-${column}-"]`);
            const checkedCount = Array.from(checkboxes).filter(cb => cb.checked).length;
            
            selectAll.checked = checkedCount === checkboxes.length;
        }

        function applyColumnFilter(column) {
            const checkboxes = document.querySelectorAll(`input[id^="filter-${column}-"]:checked`);
            const selectedValues = Array.from(checkboxes).map(cb => cb.value);
            
            const allValues = issuesData.map(i => i[column]).filter(Boolean);
            const uniqueValues = [...new Set(allValues)];
            
            if (selectedValues.length === 0 || selectedValues.length === uniqueValues.length) {
                delete activeFilters[column];
            } else {
                activeFilters[column] = selectedValues;
            }
            
            applyAllFilters();
            updateFilterStatus();
            
            if (currentDropdown) {
                currentDropdown.remove();
                currentDropdown = null;
            }
        }

        function clearColumnFilter(column) {
            delete activeFilters[column];
            applyAllFilters();
            updateFilterStatus();
            
            if (currentDropdown) {
                currentDropdown.remove();
                currentDropdown = null;
            }
        }

        function applyAllFilters() {
            let visibleCount = 0;
            
            issuesData.forEach((issue, idx) => {
                const row = document.querySelector(`tr[onclick*="handleRowClick(${idx})"]`);
                if (!row) return;
                
                let shouldShow = true;
                
                for (let column in activeFilters) {
                    if (!activeFilters[column].includes(issue[column])) {
                        shouldShow = false;
                        break;
                    }
                }
                
                row.style.display = shouldShow ? 'table-row' : 'none';
                if (shouldShow) visibleCount++;
            });
            
            document.querySelectorAll('.filterable-header').forEach(header => {
                const column = header.dataset.column;
                if (activeFilters[column]) {
                    header.classList.add('filtered');
                } else {
                    header.classList.remove('filtered');
                }
            });
            
            sendFiltersToViewer();
            
            logDebug('Showing ' + visibleCount + ' of ' + issuesData.length + ' issues');
        }

        function updateFilterStatus() {
            let statusBar = document.querySelector('.filter-status-bar');
            
            if (!statusBar) {
                statusBar = document.createElement('div');
                statusBar.className = 'filter-status-bar';
                const table = document.querySelector('.thumbnail-table');
                table.parentElement.insertBefore(statusBar, table);
            }
            
            if (Object.keys(activeFilters).length === 0) {
                statusBar.classList.remove('active');
                return;
            }
            
            statusBar.classList.add('active');
            statusBar.innerHTML = '<strong>Active Filters:</strong> ' + 
                Object.entries(activeFilters).map(([column, values]) => {
                    const label = column.replace('_', ' ').replace(/\b\w/g, l => l.toUpperCase());
                    return `<span class="filter-tag">${label}: ${values.join(', ')} <span class="remove" onclick="clearColumnFilter('${column}')">×</span></span>`;
                }).join('');
        }

        function sendFiltersToViewer() {
            const message = {
                type: 'FILTER_ISSUES',
                filters: activeFilters
            };
            
            try {
                parent.postMessage(message, '*');
                window.top.postMessage(message, '*');
                logDebug('Filters sent');
            } catch(e) {
                logDebug('Could not send filters');
            }
        }
        
        // ========== COLUMN RESIZING ==========
        function makeColumnsResizable() {
            const table = document.querySelector('.thumbnail-table');
            const cols = table.querySelectorAll('th');
            
            cols.forEach((col, index) => {
                const resizer = document.createElement('div');
                resizer.style.position = 'absolute';
                resizer.style.top = '0';
                resizer.style.right = '0';
                resizer.style.bottom = '0';
                resizer.style.width = '5px';
                resizer.style.cursor = 'col-resize';
                resizer.style.userSelect = 'none';
                resizer.style.zIndex = '1';
                
                resizer.addEventListener('mouseenter', () => {
                    resizer.style.background = 'rgba(102, 126, 234, 0.5)';
                });
                
                resizer.addEventListener('mouseleave', () => {
                    resizer.style.background = 'transparent';
                });
                
                resizer.addEventListener('mousedown', (e) => {
                    e.preventDefault();
                    const startX = e.pageX;
                    const startWidth = col.offsetWidth;
                    
                    function onMouseMove(e) {
                        const newWidth = startWidth + (e.pageX - startX);
                        if (newWidth > 50) {
                            col.style.width = newWidth + 'px';
                        }
                    }
                    
                    function onMouseUp() {
                        document.removeEventListener('mousemove', onMouseMove);
                        document.removeEventListener('mouseup', onMouseUp);
                    }
                    
                    document.addEventListener('mousemove', onMouseMove);
                    document.addEventListener('mouseup', onMouseUp);
                });
                
                col.style.position = 'relative';
                col.appendChild(resizer);
            });
        }
        
        // ========== EXCEL EXPORT FUNCTION ==========
        
        // Export to Excel WITH thumbnails (HTML method - Excel may have display issues with large images)
        function exportToExcel() {
            try {
                console.log('📥 Exporting with thumbnails...');
                
                let html = '<html xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:x="urn:schemas-microsoft-com:office:excel">';
                html += '<head><meta charset="utf-8"><style>';
                html += 'table { border-collapse: collapse; } ';
                html += 'th, td { border: 1px solid black; } ';
                html += 'th { background: #004E43; color: white; padding: 8px; font-weight: bold; } ';
                html += 'td { padding: 4px; vertical-align: middle; } ';
                html += 'img { display: block; } ';
                html += '</style></head><body><table>';
                
                // Header
                html += '<tr><th width="144">Thumbnail</th><th width="80">ID</th><th width="200">Title</th>';
                html += '<th width="80">Status</th><th width="80">Severity</th><th width="120">Assigned</th>';
                html += '<th width="200">Comments</th><th width="120">Comments By</th>';
                html += '<th width="60">X</th><th width="60">Y</th><th width="60">Z</th><th width="150">Viewable</th></tr>';
                
                // Rows - limit image size for Excel compatibility
                issuesData.forEach(issue => {
                    html += '<tr height="108">';
                    html += '<td align="center" style="padding:2px;">';
                    
                    if (issue.thumbnail && issue.thumbnail.length > 0) {
                        // For Excel compatibility, we keep the image but Excel may still have issues
                        html += '<img src="' + issue.thumbnail + '" width="192" height="144"/>';
                    } else {
                        html += 'No Image';
                    }
                    
                    html += '</td>';
                    html += '<td align="center"><b>' + (issue.display_id || '') + '</b></td>';
                    html += '<td>' + (issue.title || '') + '</td>';
                    html += '<td align="center">' + (issue.status || '') + '</td>';
                    html += '<td align="center">' + (issue.severity || '') + '</td>';
                    html += '<td>' + (issue.assigned_to || '') + '</td>';
                    html += '<td>' + (issue.comment_1 || '') + '</td>';
                    html += '<td>' + (issue.comment_1_by || '') + '</td>';
                    html += '<td align="right">' + (issue.pin_x || '') + '</td>';
                    html += '<td align="right">' + (issue.pin_y || '') + '</td>';
                    html += '<td align="right">' + (issue.pin_z || '') + '</td>';
                    html += '<td>' + (issue.viewable_name || '') + '</td>';
                    html += '</tr>';
                });
                
                html += '</table>';
                html += '<p style="margin-top:10px;"><b>Note:</b> Large images may not display correctly in Excel. ';
                html += 'If images show "cannot be displayed", try:</p>';
                html += '<ul><li>Opening file in Excel and enabling editing</li>';
                html += '<li>Using Excel Online (better base64 support)</li>';
                html += '<li>Viewing the online table instead</li></ul>';
                html += '</body></html>';
                
                const blob = new Blob([html], { type: 'application/vnd.ms-excel' });
                const url = URL.createObjectURL(blob);
                const link = document.createElement('a');
                link.href = url;
                link.download = 'Issues_With_Images_' + new Date().toISOString().split('T')[0] + '.xls';
                document.body.appendChild(link);
                link.click();
                document.body.removeChild(link);
                URL.revokeObjectURL(url);
                
                alert('✅ Exported!\\n\\n⚠️ Note: Excel has limitations with large embedded images.\\nIf images don\\'t show, click Enable Editing in Excel.');
            } catch (e) {
                console.error('Export error:', e);
                alert('Export failed: ' + e.message);
            }
        }
        
      // ========== OPEN IN EXTERNAL BROWSER ==========
        function openInBrowser() {
            // Get the current URL
            const currentUrl = window.location.href;
            
            console.log('🌐 Opening in external browser:', currentUrl);
            
            // Create a temporary link to open in new window
            const link = document.createElement('a');
            link.href = currentUrl;
            link.target = '_blank';
            link.rel = 'noopener noreferrer';
            
            // Try to click it
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
            
            // Show confirmation
            alert('✅ Opening in external browser...\\n\\nExport buttons will be available there!');
        }



        // Export to CSV (no images, but all data - opens perfectly in Excel)
        function exportToCSV() {
            try {
                let csv = 'Issue ID,Title,Status,Severity,Assigned To,Comments,Comments By,Pin X,Pin Y,Pin Z,Viewable Name\\n';
                
                issuesData.forEach(issue => {
                    csv += [
                        issue.display_id || '',
                        '"' + (issue.title || '').replace(/"/g, '""') + '"',
                        issue.status || '',
                        issue.severity || '',
                        issue.assigned_to || '',
                        '"' + (issue.comment_1 || '').replace(/"/g, '""') + '"',
                        issue.comment_1_by || '',
                        issue.pin_x || '',
                        issue.pin_y || '',
                        issue.pin_z || '',
                        issue.viewable_name || ''
                    ].join(',') + '\\n';
                });
                
                const blob = new Blob([csv], { type: 'text/csv' });
                const url = URL.createObjectURL(blob);
                const link = document.createElement('a');
                link.href = url;
                link.download = 'Issues_' + new Date().toISOString().split('T')[0] + '.csv';
                document.body.appendChild(link);
                link.click();
                document.body.removeChild(link);
                URL.revokeObjectURL(url);
                
                alert('✅ CSV exported with ' + issuesData.length + ' issues!\\n\\nNo images included, but opens perfectly in Excel.');
                
            } catch (e) {
                alert('CSV export failed: ' + e.message);
            }
        }
            

            
        // ========== INITIALIZE ON LOAD ==========
        window.addEventListener('load', function() {
            logDebug('Page loaded');
            
            // Load thumbnails
            issuesData.forEach((issue, idx) => {
                const img = document.getElementById('img_' + idx);
                if (img && issue.thumbnail) {
                    img.src = issue.thumbnail;
                    logDebug('Loaded img ' + idx);
                }
            });
            
            // Initialize filters
            initColumnFilters();
            logDebug('Filters initialized');
            
            // Make columns resizable
            makeColumnsResizable();
            
            logDebug('All loaded!');
        });
    </script>
</body>
</html>


        
        // ========== ROW CLICK HANDLER ==========
        function handleRowClick(idx) {
            const issue = issuesData[idx];
            if (!issue) {
                console.log('Issue not found at index:', idx);
                return;
            }
            console.log('🔵 Row clicked:', issue.display_id);
            logDebug('Clicked: ' + issue.display_id);
            sendMessageToViewer(issue);
        }
        
        function sendMessageToViewer(issue) {
            console.log('🔵 CLICK DETECTED:', issue.display_id);
            console.log('   Viewable GUID:', issue.viewable_guid);
            logDebug('Sending message...');
            
            const message = {
                type: 'LOAD_MODEL_AND_NAVIGATE',
                issue_id: issue.issue_id,
                display_id: issue.display_id,
                pin_x: issue.pin_x,
                pin_y: issue.pin_y,
                pin_z: issue.pin_z,
                title: issue.title,
                viewable_name: issue.viewable_name,
                viewable_guid: issue.viewable_guid,
                timestamp: Date.now()
            };
            
            console.log('📤 Sending message:', message);
            
            try {
                parent.postMessage(message, '*');
                console.log('✅ Sent to parent');
            } catch(e) {
                console.error('❌ Failed:', e);
            }
        }

"""
        
        return html
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return f"<html><body><h3>Error: {str(e)}</h3></body></html>", 500

@app.route('/powerbi-wrapper.html')
def powerbi_wrapper():
    """Wrapper page that contains both viewer and table - they can communicate directly"""
    return """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        html, body { 
            height: 100vh;
            overflow: hidden;
            font-family: 'Segoe UI', sans-serif;
        }
        #container {
            display: flex;
            flex-direction: column;
            height: 100vh;
        }
        #viewer-frame {
            flex: 1;
            border: none;
            width: 100%;
        }
        #table-frame {
            flex: 1;
            border: none;
            width: 100%;
            border-top: 5px solid #004E43;
        }
        #status {
            position: fixed;
            top: 10px;
            left: 50%;
            transform: translateX(-50%);
            background: rgba(102, 126, 234, 0.95);
            color: white;
            padding: 8px 16px;
            border-radius: 4px;
            font-size: 12px;
            z-index: 100000;
            display: none;
        }
        /* Filter header styles */
        .filterable-header {
            position: relative;
            cursor: pointer;
            user-select: none;
        }

        .filterable-header:hover {
            background: linear-gradient(135deg, #5568d3 0%, #6a4a9e 100%);
        }

        .filter-icon {
            font-size: 10px;
            margin-left: 5px;
            opacity: 0.7;
        }

        .filterable-header.filtered .filter-icon {
            color: #ffd700;
            opacity: 1;
        }

        /* Filter dropdown */
        .filter-dropdown {
            position: absolute;
            top: 100%;
            left: 0;
            background: white;
            border: 2px solid #667eea;
            border-radius: 4px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
            z-index: 1000;
            min-width: 200px;
            max-height: 300px;
            overflow-y: auto;
            display: none;
        }

        .filter-dropdown.active {
            display: block;
        }

        .filter-search {
            width: calc(100% - 20px);
            padding: 8px;
            margin: 10px;
            border: 1px solid #ddd;
            border-radius: 4px;
            font-size: 12px;
        }

        .filter-option {
            padding: 8px 12px;
            cursor: pointer;
            font-size: 12px;
            display: flex;
            align-items: center;
            gap: 8px;
            color: #333;  /* ← ADD THIS LINE */
        }

        .filter-option label {
            color: #333;  /* ← ADD THIS BLOCK */
            cursor: pointer;
        }

        .filter-option:hover {
            background: #f0f0f0;
        }

        .filter-option input[type="checkbox"] {
            cursor: pointer;
        }

        .filter-actions {
            padding: 10px;
            border-top: 1px solid #ddd;
            display: flex;
            gap: 10px;
            justify-content: space-between;
        }

        .filter-btn {
            padding: 6px 12px;
            border: none;
            border-radius: 4px;
            font-size: 11px;
            cursor: pointer;
            font-weight: 600;
        }

        .filter-btn-apply {
            background: #667eea;
            color: white;
        }

        .filter-btn-clear {
            background: #e0e0e0;
            color: #666;
        }

        .filter-status-bar {
            background: #f0f0f0;
            padding: 10px;
            margin-bottom: 10px;
            border-radius: 4px;
            font-size: 12px;
            display: none;
        }

        .filter-status-bar.active {
            display: block;
        }

        .filter-tag {
            display: inline-block;
            background: #667eea;
            color: white;
            padding: 4px 8px;
            border-radius: 12px;
            margin-right: 8px;
            font-size: 11px;
        }

        .filter-tag .remove {
            margin-left: 6px;
            cursor: pointer;
            font-weight: bold;
        }
    </style>
</head>
<body>
    <div id="status">Connecting...</div>
    <div id="container">
        <iframe id="viewer-frame" src="http://localhost:5000"></iframe>
        <iframe id="table-frame" src="http://localhost:5000/thumbnail-table.html"></iframe>
    </div>
    
    <script>
        const statusDiv = document.getElementById('status');
        
        function showStatus(msg) {
            statusDiv.textContent = msg;
            statusDiv.style.display = 'block';
            setTimeout(() => {
                statusDiv.style.display = 'none';
            }, 3000);
        }
        
        // Listen for messages from the table iframe
        window.addEventListener('message', function(event) {
            console.log('🔵 Wrapper received:', event.data);
            
        if (event.data && event.data.type === 'NAVIGATE_TO_ISSUE') {
                        console.log('✅ Relaying to viewer...');
                        const viewerFrame = document.getElementById('viewer-frame');
                        viewerFrame.contentWindow.postMessage(event.data, '*');
                        showStatus(`Navigating to ${event.data.display_id}`);
                    }
        if (event.data && event.data.type === 'LOAD_MODEL_AND_NAVIGATE') {
                        console.log('✅ Relaying LOAD_MODEL_AND_NAVIGATE to viewer...');
                        const viewerFrame = document.getElementById('viewer-frame');
                        viewerFrame.contentWindow.postMessage(event.data, '*');
                        showStatus(`Loading model for Issue ${event.data.display_id}`);
                    }            
                    if (event.data && event.data.type === 'FILTER_ISSUES') {
                        console.log('✅ Relaying filters to viewer...');
                        const viewerFrame = document.getElementById('viewer-frame');
                        viewerFrame.contentWindow.postMessage(event.data, '*');
                        showStatus(`Filtered: ${Object.keys(event.data.filters).length} columns`);
                    }
        });
        
        // Notify when both iframes are loaded
        let viewerLoaded = false;
        let tableLoaded = false;
        
        document.getElementById('viewer-frame').onload = function() {
            viewerLoaded = true;
            console.log('✅ Viewer loaded');
            if (tableLoaded) showStatus('✅ Ready!');
        };
        
        document.getElementById('table-frame').onload = function() {
            tableLoaded = true;
            console.log('✅ Table loaded');
            if (viewerLoaded) showStatus('✅ Ready!');
        };
        
    </script>
</body>
</html>
"""

@app.route('/api/debug/first-issue')
def debug_first_issue():
    """Debug endpoint - see what data is available"""
    try:
        if not issues_cache['data']:
            if FETCHER_AVAILABLE and fetch_all_issues:
                issues_cache['data'] = fetch_all_issues()
        
        issues = issues_cache['data'] or []
        
        if len(issues) > 0:
            first_issue = issues[0]
            return jsonify({
                'issue': first_issue,
                'available_fields': list(first_issue.keys()),
                'has_thumbnail_url': 'thumbnail_url' in first_issue,
                'has_thumbnail_base64': 'thumbnail_base64' in first_issue,
                'has_HTML_Image': 'HTML_Image' in first_issue,
            })
        
        return jsonify({'error': 'No issues found'})
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/health')
def health():
    """Health check"""
    return jsonify({
        'status': 'healthy',
        'issues_cached': bool(issues_cache['data']),
        'issues_count': len(issues_cache['data']) if issues_cache['data'] else 0,
        'fetcher_available': FETCHER_AVAILABLE
    })


@app.route('/api/refresh', methods=['GET', 'POST'])
def refresh():
    """Force refresh"""
    issues_cache['data'] = None
    issues_cache['timestamp'] = 0
    token_cache['token'] = None
    token_cache['expires_at'] = 0
    return jsonify({'success': True, 'message': 'Cache cleared'})


@app.route('/api/thumbnail/<int:issue_index>')
def get_thumbnail(issue_index):
    """Serve individual thumbnails as images"""
    try:
        if not issues_cache['data']:
            return jsonify({'error': 'No issues loaded'}), 404
        
        if issue_index >= len(issues_cache['data']):
            return jsonify({'error': 'Issue not found'}), 404
        
        issue = issues_cache['data'][issue_index]
        thumbnail = issue.get('thumbnail_base64', '')
        
        if not thumbnail:
            return jsonify({'error': 'No thumbnail'}), 404
        
        # Extract base64 data (remove data:image/png;base64, prefix)
        if 'base64,' in thumbnail:
            thumbnail = thumbnail.split('base64,')[1]
        
        # Decode base64
        import base64
        image_data = base64.b64decode(thumbnail)
        
        from flask import Response
        return Response(image_data, mimetype='image/png')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-excel-with-images')
def export_excel_with_images():
    """Create real Excel file with embedded images using openpyxl"""
    try:
        # You'll need: pip install openpyxl pillow
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        import base64
        from PIL import Image as PILImage
        
        if not issues_cache['data']:
            return jsonify({'error': 'No issues loaded'}), 404
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Issues"
        
        # Headers
        headers = ['Thumbnail', 'Issue ID', 'Title', 'Status', 'Severity', 'Assigned To', 
                   'Comments', 'Comments By', 'Pin X', 'Pin Y', 'Pin Z', 'Viewable']
        ws.append(headers)
        
        # Style header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = cell.font.copy(bold=True)
            cell.fill = cell.fill.copy(fgColor="004E43")
        
        # Set column widths
        ws.column_dimensions['A'].width = 26  # Thumbnail column
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 25
        
        # Add data
        issues_with_coords = [i for i in issues_cache['data'] 
                             if i.get('pin_x') and i.get('pin_y') and i.get('pin_z')]
        
        for idx, issue in enumerate(issues_with_coords, start=2):
            # Set row height for images
            ws.row_dimensions[idx].height = 108  # 1.5 inches
            
            # Add text data (skip thumbnail column)
            ws.cell(idx, 2, issue.get('display_id', ''))
            ws.cell(idx, 3, issue.get('title', ''))
            ws.cell(idx, 4, issue.get('status', ''))
            ws.cell(idx, 5, issue.get('severity', ''))
            ws.cell(idx, 6, issue.get('assigned_to', ''))
            ws.cell(idx, 7, issue.get('comment_1', ''))
            ws.cell(idx, 8, issue.get('comment_1_by', ''))
            ws.cell(idx, 9, issue.get('pin_x', ''))
            ws.cell(idx, 10, issue.get('pin_y', ''))
            ws.cell(idx, 11, issue.get('pin_z', ''))
            ws.cell(idx, 12, issue.get('viewable_name', ''))
            
            # Add thumbnail image
            thumbnail_base64 = issue.get('thumbnail_base64', '')
            if thumbnail_base64 and len(thumbnail_base64) > 0:
                try:
                    # Remove data URI prefix if present
                    if 'base64,' in thumbnail_base64:
                        thumbnail_base64 = thumbnail_base64.split('base64,')[1]
                    
                    # Decode base64
                    img_data = base64.b64decode(thumbnail_base64)
                    img_stream = BytesIO(img_data)
                    
                    # Resize image if too large
                    pil_img = PILImage.open(img_stream)
                    pil_img.thumbnail((192, 144), PILImage.Resampling.LANCZOS)
                    
                    # Save to new stream
                    output_stream = BytesIO()
                    pil_img.save(output_stream, format='PNG')
                    output_stream.seek(0)
                    
                    # Add to Excel
                    xl_img = XLImage(output_stream)
                    xl_img.width = 192  # 2 inches at 96 DPI
                    xl_img.height = 144  # 1.5 inches
                    
                    # Anchor to cell A{idx}
                    ws.add_image(xl_img, f'A{idx}')
                    
                except Exception as e:
                    print(f"Error adding image for issue {idx}: {e}")
                    ws.cell(idx, 1, "Image Error")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as download
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Issues_Export_{time.strftime("%Y%m%d")}.xlsx'
        )
        
    except ImportError:
        return jsonify({
            'error': 'Required libraries not installed',
            'message': 'Run: pip install openpyxl pillow'
        }), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("\n" + "="*70)
    print("🚀 COMPLETE SERVER - ISSUES + 3D VIEWER")
    print("="*70)
    
    print("\n📋 Status:")
    print(f"   CLIENT_ID: {'✅' if CLIENT_ID else '❌'}")
    print(f"   CLIENT_SECRET: {'✅' if CLIENT_SECRET else '❌'}")
    print(f"   HOFUF_URN: {'✅' if os.getenv('HOFUF_URN') else '❌'}")
    print(f"   SNOWDON_URN: {'✅' if os.getenv('SNOWDON_STR_URN') else '❌'}")
    print(f"   Issues Fetcher: {'✅' if FETCHER_AVAILABLE else '❌'}")
    
    # Pre-load issues and build URN mapping
    if FETCHER_AVAILABLE and fetch_all_issues:
        try:
            print("\n📥 Pre-loading issues...")
            issues_cache['data'] = fetch_all_issues()
            issues_cache['timestamp'] = time.time()
            print(f"   ✅ Loaded {len(issues_cache['data'])} issues")
            
            # Build URN mapping
            build_model_urn_mapping()
            
        except Exception as e:
            print(f"   ⚠️ Could not preload: {e}")
    
    print("\n🌐 Endpoints:")
    print("   📊 http://localhost:5000/api/issues  ← FOR POWER BI")
    print("   📺 http://localhost:5000              ← 3D Viewer")
    print("   🖼️ http://localhost:5000/thumbnail-table.html  ← Thumbnail Table")
    print("   ✅ http://localhost:5000/health       ← Status")
    
    print("\n💡 Power BI Setup:")
    print("   Get Data → Web → http://localhost:5000/api/issues")
    
    print("\n📍 All Routes:")
    for rule in app.url_map.iter_rules():
        methods = ','.join(sorted(rule.methods - {'HEAD', 'OPTIONS'}))
        print(f"   {methods:6} {rule.rule}")
    
    print("\n" + "="*70)
    print("⚡ Starting server on port 5000...")
    print("="*70 + "\n")
    
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)